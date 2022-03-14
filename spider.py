import requests
import json
import time
from openpyxl import Workbook, load_workbook
from scrapy import Selector
import re
from datetime import datetime
from http.cookiejar import MozillaCookieJar
import os
from concurrent.futures import ThreadPoolExecutor

class myntraApi():
    def __init__(self):
        self.headers = self.headers()
        cookies_path = 'myntra.com_cookies.txt'
        self.cookies = self.get_cookies(cookies_path)
        self.url = input('    Enter url (press Enter to scrape men_jeans): ')
        self.filename = input('    Enter filename (eg. men-jeans.xlsx): ')
        self.thread_limit = input('    Enter Thread Count (default is 1): ')
        self.items_to_scrape = input('    Enter no. of items to scrape (default is all): ')
        self.done = []
        if self.url == '':
            self.url = 'https://www.myntra.com/men-jeans'
        if self.filename == '':
            self.filename = 'men-jeans.xlsx'
        if self.thread_limit == '':
            self.thread_limit = 1
        if self.items_to_scrape == '':
            self.items_to_scrape = 'all'
        title = ['Crawling Time','Product Rank','Product Url','Category','Name','Brand','Product Id','Description','Seller','Average Rating','Total Rating','Total Reviews','Star1 Count','Star2 Count','Star3 Count','Star4 Count','Star5 Count','List Price','Sale Price','Product Details','Fit','Material','Product Image']
        self.page_no = 1
        self.product_count = 0
        if os.path.exists(self.filename):
            print(f'[i] {self.filename} already exists')
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
            for x in self.ws['B'][1:]:
                self.done.append(x.value)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(title)

    def headers(self):
        return {
            'authority': 'www.myntra.com',
            'x-meta-app': 'channel=web',
            'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36',
            'x-myntraweb': 'Yes',
            'content-type': 'application/json',
            'x-sec-clge-req-type': 'ajax',
            'accept': 'application/json',
            'x-myntra-app': 'deviceID=239bf692-4885-4471-b9d7-972995eda1f9;customerID=;reqChannel=web;',
            'x-requested-with': 'browser',
            'app': 'web',
            'sec-gpc': '1',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'cors',
            'sec-fetch-dest': 'empty',
            'referer': 'https://www.myntra.com/men-jeans?p=315',
            'accept-language': 'en-US,en;q=0.9',
        }

    def get_cookies(self,cookiefile):
        try:
            cj = MozillaCookieJar(cookiefile)
            cj.load()
            cookies = dict()
            for cookie in cj:
                cookies[cookie.name] = cookie.value
            print("[i] Cookies Loaded")
            return cookies
        except:
            print("[i] no file named myntra.com_cookies.txt")
            print("[i] use https://chrome.google.com/webstore/detail/get-cookiestxt/bgaddhkoddajcdgocldbbfleckgcbcid\n")
            cookies = dict()
            return cookies

    def process_item(self, item):
        # self.product_count +=1
        line = [item['Crawling Time'], item['Product Rank'] ,item['Product Url'], item['Category'], item['Name'], item['Brand'], item['Product Id'], item['Description'], item['Seller'], item['Average Rating'], item['Total Rating'], item['Total Reviews'],item['Star1 Count'], item['Star2 Count'], item['Star3 Count'], item['Star4 Count'], item['Star5 Count'], item['List Price'], item['Sale Price'], item['Product Details'], item['Fit'],item['Material'] ,item['Product Images']]
        self.ws.append(line)
        self.wb.save(self.filename)

    def page_url(self,page_no, first_product):
        if self.url[-1] == '/':
            self.url = self.url[:-1]
        keywords = self.url.split('/')[-1]
        url = f'https://www.myntra.com/gateway/v2/search/{keywords}?p={page_no}&rows=49&o={first_product}&plaEnabled=false'
        return url

    def get_requests(self,url):
        while True:
            try:
                response = requests.get(url, headers=self.headers, cookies=self.cookies)
                if response.status_code == 200:
                    break
            except:
                print('    Error in connection....Trying again')
                time.sleep(0.5)
        return response

    def product_details(self, url):
        description = ''
        fit = ''
        stars = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0
        }
        reviews_count = 0
        seller_name = ''
        material = ''
        response = self.get_requests(url)
        sel = Selector(text=response.text)
        scriptCode = sel.css('script::text').extract()
        for sc in scriptCode:
            if 'window.__myx ' in sc:
                regex = r"(?<=window.__myx = )(.*)"
                rex = re.search(regex, sc).group(0)
                json_data = json.loads(rex)
                product_details = json_data['pdpData']
                for product_detail in product_details['productDetails']:
                    title = product_detail.get('title')
                    if 'product details' in title.lower():
                        description = product_detail.get('description')
                        description = self.cleanhtml(description)
                    if 'fit' in title.lower():
                        fit = product_detail.get('description')
                        fit = self.cleanhtml(fit)
                    if 'material' in title.lower():
                        material = product_detail.get('description')
                        material = self.cleanhtml(material)
                try:
                    rating_Infos = product_details['ratings']['ratingInfo']
                    for rating_Info in rating_Infos:
                        rating = rating_Info['rating']
                        for i in range(1, 6):
                            if rating == i:
                                stars[i] = rating_Info['count']
                except:
                    pass
                try:
                    reviews_count = product_details['ratings']['reviewInfo']['reviewsCount']
                except:
                    pass
                seller_name = product_details['sellers'][0]['sellerName']  
        return description, fit, stars, reviews_count, seller_name, material

    def cleanhtml(self, raw_html):
        CLEANR = re.compile('<.*?>') 
        cleantext = re.sub(CLEANR, ' ', raw_html)
        return cleantext
 
    def get_product(self, product):
        name = product.get('productName')
        self.product_count += 1
        product['rank'] = self.product_count
        productUrl = 'https://www.myntra.com/' + product.get('landingPageUrl')
        print(f'{self.product_count}: {name}')
        if productUrl in self.done:
            return
        product_details, fit, stars, totalReviews, seller, material = self.product_details(productUrl)
        category = product.get('category')
        brand = product.get('brand')
        productId = product.get('productId')
        description = product.get('product')
        averageRating = product.get('rating')
        totalRating = product.get('ratingCount')
        star1_count = stars[1]
        star2_count = stars[2]
        star3_count = stars[3]
        star4_count = stars[4]
        star5_count = stars[5]
        list_price = product.get('mrp')
        sale_price = product.get('price')
        product_images = [image.get('src') for image in product.get('images')]
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        scraped_info = {
            'Crawling Time': dt_string,
            'Product Rank': product.get('rank'),
            'Product Url': productUrl,
            'Category': category,
            'Name': name,
            'Brand': brand,
            'Product Id': productId,
            'Description': description,
            'Seller': seller,
            'Average Rating': averageRating,
            'Total Rating': totalRating,
            'Total Reviews': totalReviews,
            'Star1 Count': star1_count,
            'Star2 Count': star2_count,
            'Star3 Count': star3_count,
            'Star4 Count': star4_count,
            'Star5 Count': star5_count,
            'List Price': list_price,
            'Sale Price': sale_price,
            'Product Details': product_details,
            'Fit': fit,
            'Material': material,
            'Product Images': ','.join(product_images)
        }
        self.process_item(scraped_info)
                    
    def get_products(self,url):
        response = self.get_requests(url)
        json_data = json.loads(response.text)
        products = json_data.get('products')
        if self.items_to_scrape == 'all':
            total_products = json_data.get('totalCount')
        else:
            total_products = int(self.items_to_scrape)
            to_scraped = total_products - self.product_count
            products = products[0:to_scraped]
        with ThreadPoolExecutor(max_workers=int(self.thread_limit)) as executor:
            executor.map(self.get_product, products)

        if self.product_count < total_products:
            more_products = True
        else:
            more_products = False
        return more_products
        
    def main(self):
        next = True
        while next:
            url = self.page_url(self.page_no, self.product_count)
            next = self.get_products(url)
            self.page_no =+ 1

if __name__=='__main__':
    mn = myntraApi()
    mn.main()